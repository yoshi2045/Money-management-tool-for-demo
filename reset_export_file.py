import shutil
import openpyxl
import consts


# 実行
if __name__ == '__main__':

    # Excelファイルのバックアップ
    shutil.copyfile(consts.EXCEL_FILE_NAME, consts.BACKUP_FILE_NAME)

    # Excelファイルを開く
    wb = openpyxl.load_workbook(consts.EXCEL_FILE_NAME)

    # Summaryシートの初期化
    sh = wb[consts.SH_SUMMARY_BANK]
    max_row = sh.max_row
    for i in range(3, max_row + 1):
        if not sh.cell(row=i, column=17).value == "○":
            # エクセル関数が埋め込まれていない行を初期化
            for m in range(4, 16): # D～O列（月の列）:
                sh.cell(row=i, column=m).value = None

    # Divideシートの初期化
    sh = wb[consts.SH_DIVIDE]
    max_row = sh.max_row
    for i in range(3, max_row + 1):
        if not sh.cell(row=i, column=17).value == "○":
            # エクセル関数が埋め込まれていない行を初期化
            for m in range(4, 16): # D～O列（月の列）:
                sh.cell(row=i, column=m).value = None

    # 各明細シートの初期化
    sheet_count = len(wb.sheetnames)
    for i in range(2, sheet_count):     # 2シート目まではスキップ
        sh = wb.worksheets[i]
        max_row = sh.max_row
        sh.delete_rows(idx=3, amount=max_row)

    # Excelファイル保存
    wb.save(consts.EXCEL_FILE_NAME)

    print("=== Reset has done! ===")
