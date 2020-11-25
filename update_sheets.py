import sys
import os
import shutil
import openpyxl
import consts


# 各明細シートからSummary、Divideシート更新
def execute(wb):
    analyze_bank_ufj_family(wb)
    analyze_bank_ufj_biz(wb)
    analyze_bank_rakuten(wb)
    analyze_bank_neo(wb)
    analyze_bank_smbc(wb)
    analyze_card_rakuten(wb)


# 「Bank_UFJ(Family)」シートの解析
def analyze_bank_ufj_family(wb):
    sh = wb[consts.SH_BANK_UFJ_FAMILY]

    # リスト初期化
    balance_list = {}   # 月末残高
    water_list = {}     # 水道代
    for i in range(1, 13):
        balance_list[i] = ""
        water_list[i] = ""

    # 明細行解析
    # A列が１行以上データがあり、かつ途中欠落していないことが前提
    for row_no in range(3, sh.max_row):     # max_row - 1 の行までループする
        str = sh.cell(row=row_no, column=2).value

        # 水道代判定
        if str.startswith("水道"):
            current_date = sh.cell(row=row_no, column=1).value
            water_list[current_date.month] = sh.cell(row=row_no, column=4).value

        # 月末残高取得
        current_date = sh.cell(row=row_no, column=1).value
        next_date = sh.cell(row=row_no + 1, column=1).value
        if row_no + 1 == sh.max_row:
            # ループの最後はnext_dateが月末残高
            # （水道代は月末最終行にならない想定）
            balance_list[next_date.month] = sh.cell(row=row_no + 1, column=6).value
        elif current_date.month < next_date.month:
            balance_list[current_date.month] = sh.cell(row=row_no, column=6).value

    # シート切り替え
    sh = wb[consts.SH_SUMMARY_BANK]
    offset = 3
    # 月末残高更新
    for k in balance_list.keys():
        sh.cell(row=8, column=offset + k).number_format = '#,##0'
        sh.cell(row=8, column=offset + k).value = balance_list[k]

    # シート切り替え
    sh = wb[consts.SH_DIVIDE]
    # 月毎料金更新
    for i in water_list:
        sh.cell(row=12, column=offset + i).number_format = '#,##0'
        sh.cell(row=12, column=offset + i).value = water_list[i]


# 「Bank_UFJ(Biz)」シートの解析
def analyze_bank_ufj_biz(wb):
    sh = wb[consts.SH_BANK_UFJ_BIZ]

    # 月末残高リスト初期化
    balance_list = {}
    for i in range(1, 13):
        balance_list[i] = ""

    # 明細行解析
    # A列が１行以上データがあり、かつ途中欠落していないことが前提
    for row_no in range(3, sh.max_row):     # max_row - 1 の行までループする
        # 月末残高取得
        current_date = sh.cell(row=row_no, column=1).value
        next_date = sh.cell(row=row_no + 1, column=1).value
        if row_no + 1 == sh.max_row:
            # ループの最後はnext_dateが月末残高
            balance_list[next_date.month] = sh.cell(row=row_no + 1, column=6).value
        elif current_date.month < next_date.month:
            balance_list[current_date.month] = sh.cell(row=row_no, column=6).value

    # シート切り替え
    sh = wb[consts.SH_SUMMARY_BANK]
    offset = 3
    # 月末残高更新
    for k in balance_list.keys():
        sh.cell(row=12, column=offset + k).number_format = '#,##0'
        sh.cell(row=12, column=offset + k).value = balance_list[k]


# 「Bank_Rakuten」シートの解析
def analyze_bank_rakuten(wb):
    sh = wb[consts.SH_BANK_RAKUTEN]

    # 月末残高リスト初期化
    balance_list = {}
    for i in range(1, 13):
        balance_list[i] = ""

    # 明細行解析
    # A列が１行以上データがあり、かつ途中欠落していないことが前提
    for row_no in range(3, sh.max_row):     # max_row - 1 の行までループする
        # 月末残高取得
        current_date = sh.cell(row=row_no, column=1).value
        next_date = sh.cell(row=row_no + 1, column=1).value
        if row_no + 1 == sh.max_row:
            # ループの最後はnext_dateが月末残高
            # 月が連続していない場合もあるので、for文で途中の月の値を埋める
            for i in range(current_date.month, next_date.month):
                balance_list[i] = sh.cell(row=row_no, column=3).value
            balance_list[next_date.month] = sh.cell(row=row_no + 1, column=3).value
        elif current_date.month < next_date.month:
            # 月が連続していない場合もあるので、for文で途中の月の値を埋める
            for i in range(current_date.month, next_date.month):
                balance_list[i] = sh.cell(row=row_no, column=3).value

    # シート切り替え
    sh = wb[consts.SH_SUMMARY_BANK]
    offset = 3
    # 月末残高更新
    for k in balance_list.keys():
        sh.cell(row=16, column=offset + k).number_format = '#,##0'
        sh.cell(row=16, column=offset + k).value = balance_list[k]


# 「Bank_NEO」シートの解析
def analyze_bank_neo(wb):
    sh = wb[consts.SH_BANK_NEO]

    # 月末残高リスト初期化
    balance_list = {}
    for i in range(1, 13):
        balance_list[i] = ""

    # 明細行解析
    # A列が１行以上データがあり、かつ途中欠落していないことが前提
    for row_no in range(3, sh.max_row):     # max_row - 1 の行までループする
        # 月末残高取得
        current_date = sh.cell(row=row_no, column=1).value
        next_date = sh.cell(row=row_no + 1, column=1).value
        if row_no + 1 == sh.max_row:
            # ループの最後はnext_dateが月末残高
            # 月が連続していない場合もあるので、for文で途中の月の値を埋める
            for i in range(current_date.month, next_date.month):
                balance_list[i] = sh.cell(row=row_no, column=5).value
            balance_list[next_date.month] = sh.cell(row=row_no + 1, column=5).value
        elif current_date.month < next_date.month:
            # 月が連続していない場合もあるので、for文で途中の月の値を埋める
            for i in range(current_date.month, next_date.month):
                balance_list[i] = sh.cell(row=row_no, column=5).value

    # シート切り替え
    sh = wb[consts.SH_SUMMARY_BANK]
    offset = 3
    # 月末残高更新
    for k in balance_list.keys():
        sh.cell(row=20, column=offset + k).number_format = '#,##0'
        sh.cell(row=20, column=offset + k).value = balance_list[k]


# 「Bank_SMBC」シートの解析
def analyze_bank_smbc(wb):
    sh = wb[consts.SH_BANK_SMBC]

    # 月末残高リスト初期化
    balance_list = {}
    for i in range(1, 13):
        balance_list[i] = ""

    # 明細行解析
    # A列が１行以上データがあり、かつ途中欠落していないことが前提
    for row_no in range(3, sh.max_row):     # max_row - 1 の行までループする
        # 月末残高取得
        current_date = sh.cell(row=row_no, column=1).value
        next_date = sh.cell(row=row_no + 1, column=1).value
        if row_no + 1 == sh.max_row:
            # ループの最後はnext_dateが月末残高
            # 月が連続していない場合もあるので、for文で途中の月の値を埋める
            for i in range(current_date.month, next_date.month):
                balance_list[i] = sh.cell(row=row_no, column=5).value
            balance_list[next_date.month] = sh.cell(row=row_no + 1, column=5).value
        elif current_date.month < next_date.month:
            # 月が連続していない場合もあるので、for文で途中の月の値を埋める
            for i in range(current_date.month, next_date.month):
                balance_list[i] = sh.cell(row=row_no, column=5).value

    # シート切り替え
    sh = wb[consts.SH_SUMMARY_BANK]
    offset = 3
    # 月末残高更新
    for k in balance_list.keys():
        sh.cell(row=24, column=offset + k).number_format = '#,##0'
        sh.cell(row=24, column=offset + k).value = balance_list[k]


# 「Card_Rakuten」シートの解析
def analyze_card_rakuten(wb):

    # リスト初期化
    gass_list = {}      # ガス代
    ep_list = {}        # 電気代
    for i in range(1, 13):
        gass_list[i] = ""
        ep_list[i] = ""

    # 集計対象年度取得
    sh = wb[consts.SH_DIVIDE]
    current_year = sh.cell(row=2, column=3).value

    # シート切り替え
    sh = wb[consts.SH_CARD_RAKUTEN]

    # 明細行解析
    # A列が１行以上データがあり、かつ途中欠落していないことが前提
    for row_no in range(3, sh.max_row + 1):     # max_row の行までループする
        str = sh.cell(row=row_no, column=2).value

        # ガス代判定
        if str.startswith("ｶﾞｽ"):
            current_date = sh.cell(row=row_no, column=1).value
            # 今年度のデータか確認
            if current_date.year == current_year:
                gass_list[current_date.month] = sh.cell(row=row_no, column=5).value

        # 電気代判定
        elif str.startswith("東京電力") or str.startswith("ｼ-ﾃﾞｲ-ｴﾅｼﾞ-ﾀﾞｲﾚｸﾄ"):
            current_date = sh.cell(row=row_no, column=1).value
            # 今年度のデータか確認
            if current_date.year == current_year:
                ep_list[current_date.month] = sh.cell(row=row_no, column=5).value

    # シート切り替え
    sh = wb[consts.SH_DIVIDE]
    offset = 3
    # 月毎料金更新
    for i in range(1, 13):
        # ガス代
        sh.cell(row=4, column=offset + i).number_format = '#,##0'
        sh.cell(row=4, column=offset + i).value = gass_list[i]

        # 電気代
        sh.cell(row=8, column=offset + i).number_format = '#,##0'
        sh.cell(row=8, column=offset + i).value = ep_list[i]
