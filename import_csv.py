import sys
import os
import shutil
import pathlib
import csv
import openpyxl
import datetime
import mmtUtil
import consts


# Excelの各シートへ書き込み
def execute(wb):
    # インポート用データフォルダのパスオブジェクトを取得
    data_path = pathlib.Path(consts.DATA_PATH)

    # インポートと出力
    for pass_obj in data_path.iterdir():
        # csvファイル以外はスキップ
        if not pass_obj.match('*.csv'):  continue

        # データファイルの前方一致で書き込み先シートを振り分け
        if pass_obj.name.startswith("12"):
            sh = wb[consts.SH_BANK_UFJ_BIZ]
            import_to_bank_ufj(pass_obj, sh)
        elif pass_obj.name.startswith("76"):
            sh = wb[consts.SH_BANK_UFJ_FAMILY]
            import_to_bank_ufj(pass_obj, sh)
        elif pass_obj.name.startswith("RB"):
            sh = wb[consts.SH_BANK_RAKUTEN]
            import_to_bank_rakuten(pass_obj, sh)
        elif pass_obj.name.startswith("nyushukinmeisai"):
            sh = wb[consts.SH_BANK_NEO]
            import_to_bank_neo(pass_obj, sh)
        elif pass_obj.name.startswith("meisai"):
            sh = wb[consts.SH_BANK_SMBC]
            import_to_bank_smbc(pass_obj, sh)
        elif pass_obj.name.startswith("enavi"):
            sh = wb[consts.SH_CARD_RAKUTEN]
            import_to_card_rakuten(pass_obj, sh)
        elif pass_obj.name.startswith("ご利用明細"):
            sh = wb[consts.SH_CARD_VIEW]
            import_to_card_view(pass_obj, sh)
        elif pass_obj.name.startswith("SAISON"):
            sh = wb[consts.SH_CARD_SAISON]
            import_to_card_saison(pass_obj, sh)


# UFJ銀行のcsvデータを書き込み
def import_to_bank_ufj(pass_obj, sh):
    # csvファイルを読み込み、出力先シートに追加
    with open(pass_obj, 'r', encoding=mmtUtil.get_encoding(pass_obj)) as f:
        reader = csv.reader(f)
        next(f)  # 1行目を読み飛ばし
        csv_data = [row for row in reader]  # リスト内包表記
        max_row = sh.max_row
        for row_no, row in enumerate(csv_data, max_row + 1):
            # 日付
            dttm = datetime.datetime.strptime(row[0], '%Y/%m/%d')
            sh.cell(row=row_no, column=1).value = dttm
            sh.cell(row=row_no, column=1).number_format = 'mm-dd-yy'

            # 摘要
            sh.cell(row=row_no, column=2, value=row[1])
            # 摘要内容
            sh.cell(row=row_no, column=3, value=row[2])

            # 支払金額
            if not row[3] == '':
                sh.cell(row=row_no, column=4).number_format = '#,##0'
                sh.cell(row=row_no, column=4).value = int(row[3].replace(',', ''))

            # 預かり金額
            if not row[4] == '':
                sh.cell(row=row_no, column=5).number_format = '#,##0'
                sh.cell(row=row_no, column=5).value = int(row[4].replace(',', ''))

            # 差引残高
            if not row[5] == '':
                sh.cell(row=row_no, column=6).number_format = '#,##0'
                sh.cell(row=row_no, column=6).value = int(row[5].replace(',', ''))

            # メモ
            sh.cell(row=row_no, column=7, value=row[6])
            # 未資金化区分
            sh.cell(row=row_no, column=8, value=row[7])
            # 入払区分
            sh.cell(row=row_no, column=9, value=row[8])


# 楽天銀行のcsvデータを書き込み
def import_to_bank_rakuten(pass_obj, sh):
    # csvファイルを読み込み、出力先シートに追加
    with open(pass_obj, 'r', encoding=mmtUtil.get_encoding(pass_obj)) as f:
        reader = csv.reader(f)
        next(f)  # 1行目を読み飛ばし
        csv_data = [row for row in reader]  # リスト内包表記
        max_row = sh.max_row
        for row_no, row in enumerate(csv_data, max_row + 1):
            # 取引日
            dttm = datetime.datetime.strptime(row[0], '%Y%m%d')
            sh.cell(row=row_no, column=1).value = dttm
            sh.cell(row=row_no, column=1).number_format = 'mm-dd-yy'

            # 入出金（円）
            if not row[1] == '':
                sh.cell(row=row_no, column=2).number_format = '#,##0'
                sh.cell(row=row_no, column=2).value = int(row[1])

            # 残高（円）
            if not row[2] == '':
                sh.cell(row=row_no, column=3).number_format = '#,##0'
                sh.cell(row=row_no, column=3).value = int(row[2])

            # 入出金先内容
            sh.cell(row=row_no, column=4, value=row[3])


# NEOBANK（旧SBI銀行）のcsvデータを書き込み
def import_to_bank_neo(pass_obj, sh):
    # csvファイルを読み込み、出力先シートに追加
    with open(pass_obj, 'r', encoding=mmtUtil.get_encoding(pass_obj)) as f:
        reader = csv.reader(f)
        next(f)  # 1行目を読み飛ばし
        csv_data = [row for row in reader]  # リスト内包表記
        rvs_csv_data = reversed(csv_data)   # データが降順になっているので、読み込み後に逆順にする
        max_row = sh.max_row
        for row_no, row in enumerate(rvs_csv_data, max_row + 1):
            # 日付
            dttm = datetime.datetime.strptime(row[0], '%Y/%m/%d')
            sh.cell(row=row_no, column=1).value = dttm
            sh.cell(row=row_no, column=1).number_format = 'mm-dd-yy'

            # 内容
            sh.cell(row=row_no, column=2, value=row[1])

            # 出金金額(円)
            if not row[2] == '':
                sh.cell(row=row_no, column=3).number_format = '#,##0'
                sh.cell(row=row_no, column=3).value = int(row[2].replace(',', ''))

            # 入金金額(円)
            if not row[3] == '':
                sh.cell(row=row_no, column=4).number_format = '#,##0'
                sh.cell(row=row_no, column=4).value = int(row[3].replace(',', ''))

            # 残高(円)
            if not row[4] == '':
                sh.cell(row=row_no, column=5).number_format = '#,##0'
                sh.cell(row=row_no, column=5).value = int(row[4].replace(',', ''))

            # メモ
            sh.cell(row=row_no, column=6, value=row[5])


# SMBC（三井住友銀行）のcsvデータを書き込み
def import_to_bank_smbc(pass_obj, sh):
    # csvファイルを読み込み、出力先シートに追加
    with open(pass_obj, 'r', encoding=mmtUtil.get_encoding(pass_obj)) as f:
        reader = csv.reader(f)
        next(f)  # 1行目を読み飛ばし
        csv_data = [row for row in reader]  # リスト内包表記
        max_row = sh.max_row
        for row_no, row in enumerate(csv_data, max_row + 1):
            # SMBCのエクスポートcsvデータは、
            # 最終行に空行（改行のみの行）があるので読み飛ばす必要あり
            # 最終行なら終了
            if not row:
                break

            # 年月日
            str = mmtUtil.convert_sbmc_wareki_to_ad(row[0])
            dttm = datetime.datetime.strptime(str, '%Y/%m/%d')
            sh.cell(row=row_no, column=1).value = dttm
            sh.cell(row=row_no, column=1).number_format = 'mm-dd-yy'

            # お引出し
            if not row[1] == '':
                sh.cell(row=row_no, column=2).number_format = '#,##0'
                sh.cell(row=row_no, column=2).value = int(row[1])

            # お預入れ
            if not row[2] == '':
                sh.cell(row=row_no, column=3).number_format = '#,##0'
                sh.cell(row=row_no, column=3).value = int(row[2])

            # お取り扱い内容
            sh.cell(row=row_no, column=4, value=row[3])

            # お預入れ
            if not row[4] == '':
                sh.cell(row=row_no, column=5).number_format = '#,##0'
                sh.cell(row=row_no, column=5).value = int(row[4])


# 楽天カードのcsvデータを書き込み
def import_to_card_rakuten(pass_obj, sh):
    # csvファイルを読み込み、出力先シートに追加
    with open(pass_obj, 'r', encoding=mmtUtil.get_encoding(pass_obj)) as f:
        reader = csv.reader(f)
        next(f)  # 1行目を読み飛ばし
        csv_data = [row for row in reader]  # リスト内包表記
        max_row = sh.max_row
        idx_csv_data = 0
        offset = 0
        for row_no, row in enumerate(csv_data, max_row + 1):
            if row[0] == '':
                # 先頭列が空欄の場合はスキップ
                idx_csv_data += 1
                offset += 1
                continue

            # 利用店名・商品名が複数行のデータを検出
            shop_name = row[1]
            if idx_csv_data < len(csv_data) - 1:
                if csv_data[idx_csv_data + 1][0] == '' and not csv_data[idx_csv_data + 1][1] == '':
                    shop_name = csv_data[idx_csv_data][1] + '\r\n' + csv_data[idx_csv_data + 1][1]

            new_row_no = row_no - offset

            # 利用日
            dttm = datetime.datetime.strptime(row[0], '%Y/%m/%d')
            sh.cell(row=new_row_no, column=1).value = dttm
            sh.cell(row=new_row_no, column=1).number_format = 'mm-dd-yy'

            # 利用店名・商品名
            sh.cell(row=new_row_no, column=2).alignment = openpyxl.styles.Alignment(wrapText=True)
            sh.cell(row=new_row_no, column=2, value=shop_name)

            # 利用者
            sh.cell(row=new_row_no, column=3, value=row[2])

            # 支払方法
            sh.cell(row=new_row_no, column=4, value=row[3])

            # 利用金額
            if not row[4] == '':
                sh.cell(row=new_row_no, column=5).number_format = '#,##0'
                sh.cell(row=new_row_no, column=5).value = int(row[4])

            # 支払手数料
            if not row[5] == '':
                sh.cell(row=new_row_no, column=6).number_format = '#,##0'
                sh.cell(row=new_row_no, column=6).value = int(row[5])

            # 支払総額
            if not row[6] == '':
                sh.cell(row=new_row_no, column=7).number_format = '#,##0'
                sh.cell(row=new_row_no, column=7).value = int(row[6])

            # 当月支払金額
            if not row[7] == '':
                sh.cell(row=new_row_no, column=8).number_format = '#,##0'
                sh.cell(row=new_row_no, column=8).value = int(row[7])

            # 翌月繰越残高
            if not row[8] == '':
                sh.cell(row=new_row_no, column=9).number_format = '#,##0'
                sh.cell(row=new_row_no, column=9).value = int(row[8])

            # 新規サイン
            sh.cell(row=new_row_no, column=10, value=row[9])

            idx_csv_data += 1


# Viewカードのcsvデータを書き込み
def import_to_card_view(pass_obj, sh):
    # csvファイルを読み込み、出力先シートに追加
    with open(pass_obj, 'r', encoding=mmtUtil.get_encoding(pass_obj)) as f:
        reader = csv.reader(f)
        # 明細行まで読み飛ばし(上の行のすぐ後でないと動作しないので注意)
        for i in range(7):
            next(f)

        csv_data = [row for row in reader]  # リスト内包表記
        max_row = sh.max_row

        for row_no, row in enumerate(csv_data, max_row + 1):
            # 明細行下部にある、払戻に関するデータは読み込まない
            if row[0] == "" and row[1] == "ご入金等による充当額":
                break

            # ご利用年月日
            dttm = datetime.datetime.strptime(row[0], '%Y/%m/%d')
            sh.cell(row=row_no, column=1).value = dttm
            sh.cell(row=row_no, column=1).number_format = 'mm-dd-yy'

            # ご利用箇所
            sh.cell(row=row_no, column=2, value=row[1])

            # ご利用額
            if not row[2] == '':
                sh.cell(row=row_no, column=3).number_format = '#,##0'
                sh.cell(row=row_no, column=3).value = int(row[2].replace(',', ''))

            # 払戻額
            if not row[3] == '':
                sh.cell(row=row_no, column=4).number_format = '#,##0'
                sh.cell(row=row_no, column=4).value = int(row[3].replace(',', ''))

            # ご請求額（うち手数料・利息）
            if not row[4] == '':
                sh.cell(row=row_no, column=5).number_format = '#,##0'
                sh.cell(row=row_no, column=5).value = int(row[4].replace(',', ''))

            # 支払区分（回数）
            sh.cell(row=row_no, column=6, value=row[5])

            # 今回回数
            sh.cell(row=row_no, column=7, value=row[6])

            # 今回ご請求額・弁済金（うち手数料・利息）
            if not row[7] == '':
                sh.cell(row=row_no, column=8).number_format = '#,##0'
                sh.cell(row=row_no, column=8).value = int(row[7].replace(',', ''))

            # 現地通貨額
            sh.cell(row=row_no, column=9, value=row[8])

            # 通貨略称
            sh.cell(row=row_no, column=10, value=row[9])

            # 換算レート
            sh.cell(row=row_no, column=11, value=row[10])


# SAISONカードのcsvデータを書き込み
def import_to_card_saison(pass_obj, sh):
    # csvファイルを読み込み、出力先シートに追加
    with open(pass_obj, 'r', encoding=mmtUtil.get_encoding(pass_obj)) as f:
        reader = csv.reader(f)
        # 明細行まで読み飛ばし(上の行のすぐ後でないと動作しないので注意)
        for i in range(5):
            next(f)

        csv_data = [row for row in reader]  # リスト内包表記
        max_row = sh.max_row

        for row_no, row in enumerate(csv_data, max_row + 1):
            # 利用日
            dttm = datetime.datetime.strptime(row[0], '%Y/%m/%d')
            sh.cell(row=row_no, column=1).value = dttm
            sh.cell(row=row_no, column=1).number_format = 'mm-dd-yy'

            # ご利用店名及び商品名
            sh.cell(row=row_no, column=2, value=row[1])

            # 本人・家族区分
            sh.cell(row=row_no, column=3, value=row[2])

            # 支払区分名称
            sh.cell(row=row_no, column=4, value=row[3])

            # 締前入金区分
            sh.cell(row=row_no, column=5, value=row[4])

            # 利用金額
            if not row[5] == '':
                sh.cell(row=row_no, column=6).number_format = '#,##0'
                sh.cell(row=row_no, column=6).value = int(row[5].replace(',', ''))

            # 備考
            sh.cell(row=row_no, column=7, value=row[6])
